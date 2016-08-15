from openpyxl import Workbook, load_workbook
from sklearn import datasets, linear_model


_NH_FileLocal = './Data/Samples/NH.xlsx'
_NO_FileLocal = './Data/Samples/NO.xlsx'

TAG_NH = 1
TAG_NO = 2


def __Load_NH_Data() :
    workBook = load_workbook(filename = _NH_FileLocal, read_only = True)
    sheet = workBook['Sheet1']

    x_Parameter = []
    y_Parameter = []

    for cell in sheet.columns[0] :
        x_Parameter.append(cell.value)
    del x_Parameter[0]

    for cell in sheet.columns[1] :
        y_Parameter.append(cell.value)
    del y_Parameter[0]

    return x_Parameter, y_Parameter


def __Load_NO_Data() :
    workBook = load_workbook(filename = _NO_FileLocal, read_only = True)
    sheet = workBook['Sheet1']

    x_Parameter = []
    y_Parameter = []

    for cell in sheet.columns[0] :
        x_Parameter.append(cell.value)
    del x_Parameter[0]

    for cell in sheet.columns[1] :
        y_Parameter.append(cell.value)
    del y_Parameter[0]

    return x_Parameter, y_Parameter


def Get_PredictResult(value = 0.0, tag = 1) :
    linear = linear_model.LinearRegression()

    if tag is 1 :
        linear.fit(__Load_NH_Data())
        result = linear.predict(value)
        return result
    elif tag is 2 :
        linear.fit(__Load_NO_Data())
        result = linear.predict(value)
        return result
