from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
import NetTime

_fileLocal = './Data/Result/'


def __Load_WorkBook() :
    realTime = NetTime.NetTime()
    realTime.UpdateTime()

    fileName = realTime.Date + '.xlsx'

    try :
        workBook = load_workbook(filename = _fileLocal + fileName)
    except :
        workBook = Workbook()
        sheet = workBook.active
        sheet.append(['Time', 'NO', 'NH'])
        workBook.save(filename = _fileLocal + fileName)
        workBook = load_workbook(filename = _fileLocal + fileName)

    return workBook


def Write_NewResult(no = 0.0, nh = 0.0) :
    realTime = NetTime.NetTime()
    realTime.UpdateTime()

    workBook = __Load_WorkBook()

    sheet = workBook.active

    if no < 0 :
        cell = sheet.cell(column = 2, 
                          row = sheet.max_row)
        no = float(cell.value)

    if nh < 0 :
        cell = sheet.cell(column = 3, 
                          row = sheet.max_row)
        nh = float(cell.value)

    sheet.append([realTime.Time, no, nh])

    try :
        chartSheet = workBook.get_sheet_by_name('Analyze')
    except :
        chartSheet = workBook.create_chartsheet(title = 'Analyze')

    chart = LineChart()
    chart.title = 'Analyze'
    chart.style = 10

    chart.y_axis.title = 'Value'
    chart.x_axis.title = 'Time'

    area = Reference(sheet, 
                     min_col = 2, max_col = sheet.max_column, 
                     min_row = 1, max_row = sheet.max_row)
    chart.add_data(area, titles_from_data = True)

    date = Reference(sheet, 
                     min_col = 1,
                     min_row = 2, max_row = sheet.max_row)
    chart.set_categories(date)

    chartSheet.add_chart(chart)

    fileName = realTime.Date + '.xlsx'
    workBook.save(filename = _fileLocal + fileName)

